#!/usr/bin/env python3
"""
generate_report.py — Membaca LAPORAN KAS SD6.xlsx dan menghasilkan data.js
Jalankan: python generate_report.py
"""

import json
import re
import sys
import warnings
from collections import defaultdict
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Pastikan openpyxl tersedia ─────────────────────────────────────
try:
    import openpyxl
except ImportError:
    import subprocess
    print("Menginstal openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Path ───────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "LAPORAN KAS SD6.xlsx"
OUTPUT_JS  = BASE_DIR / "data.js"

# ── Urutan bulan (Juli–Juni) ───────────────────────────────────────
BULAN_ORDER = [
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
]
BULAN_LABEL = {
    "Juli":      "Juli 2025",
    "Agustus":   "Agustus 2025",
    "September": "September 2025",
    "Oktober":   "Oktober 2025",
    "November":  "November 2025",
    "Desember":  "Desember 2025",
    "Januari":   "Januari 2026",
    "Februari":  "Februari 2026",
    "Maret":     "Maret 2026",
    "April":     "April 2026",
    "Mei":       "Mei 2026",
    "Juni":      "Juni 2026",
}


# ══════════════════════════════════════════════════════════════════
# 1. BACA DATA DARI EXCEL
# ══════════════════════════════════════════════════════════════════

def load_cashflow(wb):
    """Baca semua transaksi dari sheet 'CashFlow (2)'."""
    ws = wb["CashFlow (2)"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        no, tanggal, bulan, keterangan, arus_kas, kategori, sub, value, saldo, siswa = row[:10]
        if value is None or arus_kas not in ("KAS MASUK", "KAS KELUAR"):
            continue
        rows.append({
            "no":         no,
            "tanggal":    tanggal,
            "bulan":      bulan or "",
            "keterangan": keterangan or "",
            "arus_kas":   arus_kas,
            "kategori":   kategori or "",
            "sub":        sub or "",
            "value":      float(value),
            "saldo":      float(saldo) if saldo is not None else 0.0,
            "siswa":      str(siswa or "").strip(),
        })
    return rows


def load_tawaun_siswa(wb):
    """Baca donasi ta'awun per siswa dari sheet pivot 'Ta'awun'."""
    ws = wb["Ta'awun"]
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        nama, nilai = row[0], row[1]
        if nama and nama != "Grand Total" and nilai:
            result.append({"nama": nama, "nilai": int(nilai)})
    return sorted(result, key=lambda x: -x["nilai"])


def load_rencana(wb):
    """Baca rencana pengeluaran dari sheet 'Rencana Pengeluaran'."""
    ws = wb["Rencana Pengeluaran"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        no = row[0]
        if not isinstance(no, int):
            continue
        bulan_raw = row[2] if len(row) > 2 else None
        bulan_str = str(bulan_raw) if bulan_raw else ""
        keterangan = row[3]
        kategori   = row[4]
        sub        = row[5]
        kas_keluar = row[7]
        if kas_keluar:
            result.append({
                "no":          no,
                "bulan":       bulan_str,
                "keterangan":  keterangan or "",
                "kategori":    kategori or "",
                "sub":         sub or "",
                "nilai":       float(kas_keluar),
            })
    return result


def load_siswa(wb):
    """Baca daftar siswa dari sheet 'Tabungan Anak' (34 siswa aktif)."""
    ws = wb["Tabungan Anak"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        no   = row[0]
        nama = row[1]
        if isinstance(no, int) and nama:
            result.append(nama.strip())
    return result


def load_total_tabungan(wb):
    """Jumlah TOTAL SALDO (kolom Y) dari sheet 'Tabungan Anak' — basis saldo yang akan dikembalikan ke siswa."""
    ws = wb["Tabungan Anak"]
    total = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], int):
            continue
        val = row[24]  # kolom Y: TOTAL SALDO
        if val is not None:
            total += float(val)
    return total


# ══════════════════════════════════════════════════════════════════
# 2. KOMPUTASI DATA
# ══════════════════════════════════════════════════════════════════

def compute_summary(txns):
    total_masuk  = sum(t["value"] for t in txns if t["arus_kas"] == "KAS MASUK")
    total_keluar = sum(t["value"] for t in txns if t["arus_kas"] == "KAS KELUAR")
    saldo_akhir  = total_masuk - total_keluar

    # Saldo tertinggi = saldo akhir bulan tertinggi (bukan running balance mid-bulan)
    month_end = defaultdict(float)
    for t in txns:
        if t["bulan"]:
            month_end[t["bulan"]] = t["saldo"]
    if month_end:
        peak_bulan = max(month_end, key=lambda b: month_end[b])
        saldo_max  = month_end[peak_bulan]
        bulan_peak = BULAN_LABEL.get(peak_bulan, peak_bulan)
    else:
        saldo_max, bulan_peak = 0, ""

    return {
        "totalMasuk":     round(total_masuk),
        "totalKeluar":    round(total_keluar),
        "saldoAkhir":     round(saldo_akhir),
        "saldoTertinggi": round(saldo_max),
        "bulanTertinggi": bulan_peak,
        "jumlahTransaksi": len(txns),
    }


def compute_rekap(txns):
    categories = ["Kas", "Kegiatan", "Ta'awun"]
    cat_labels  = {"Kas": "💰 Kas", "Kegiatan": "🎒 Kegiatan", "Ta'awun": "🤝 Ta'awun"}
    cat_desc    = {
        "Kas":      "Iuran, operasional, konsumsi, administrasi",
        "Kegiatan": "Outing, penelitian, OASIXTH, graduation",
        "Ta'awun":  "Donasi sosial, santunan, THR guru",
    }
    result = []
    for cat in categories:
        masuk  = sum(t["value"] for t in txns if t["kategori"] == cat and t["arus_kas"] == "KAS MASUK")
        keluar = sum(t["value"] for t in txns if t["kategori"] == cat and t["arus_kas"] == "KAS KELUAR")
        result.append({
            "kategori": cat,
            "label":    cat_labels[cat],
            "desc":     cat_desc[cat],
            "masuk":    round(masuk),
            "keluar":   round(keluar),
            "saldo":    round(masuk - keluar),
        })
    return result


def compute_monthly(txns):
    """Saldo akhir bulan + net movement per bulan."""
    month_txns = defaultdict(list)
    for t in txns:
        if t["bulan"] in BULAN_ORDER:
            month_txns[t["bulan"]].append(t)

    result     = []
    prev_saldo = 0
    for bulan in BULAN_ORDER:
        if bulan not in month_txns:
            continue
        mt        = month_txns[bulan]
        # Saldo akhir = saldo transaksi terakhir bulan ini
        saldo_end = mt[-1]["saldo"]
        net       = round(saldo_end - prev_saldo)

        # Highlights: keterangan dari 3 transaksi terbesar nilainya
        top3       = sorted(mt, key=lambda t: t["value"], reverse=True)
        highlights = []
        seen       = set()
        for t in top3:
            k = t["keterangan"].strip()
            if k and k not in seen and len(highlights) < 3:
                # Potong teks panjang
                highlights.append(k[:60] + ("…" if len(k) > 60 else ""))
                seen.add(k)

        result.append({
            "bulan":     BULAN_LABEL.get(bulan, bulan),
            "saldoAkhir": round(saldo_end),
            "net":        net,
            "highlight":  ", ".join(highlights),
        })
        prev_saldo = saldo_end

    return result


def compute_kas_breakdown(txns):
    """Pecahan sub-kategori untuk kategori Kas."""
    kas = [t for t in txns if t["kategori"] == "Kas"]

    # Display name mapping untuk KAS MASUK (sub di Excel → label di laporan)
    KAS_MASUK_LABEL = {
        "Iuran Kas": "Iuran Kas",
        "Lainnya":   "Saldo Pindahan",   # saldo SD5 masuk sebagai "Lainnya" di Excel
        "Donasi":    "Donasi Balik",      # pengembalian transportasi
    }
    KAS_MASUK_DESC = {
        "Iuran Kas": "34 siswa lunas (Sem.1 + Sem.2)",
        "Lainnya":   "Sisa kas SD5 + sisa outing TNWK",
        "Donasi":    "Pengembalian transportasi dikembalikan ke kas",
    }
    KAS_KELUAR_DESC = {
        "Lainnya":      "HUT RI Paniisan, dekor tumpeng, subsidi kegiatan, dll",
        "Operasional":  "Air galon, kebersihan kelas (rutin setiap bulan)",
        "Konsumsi":     "Snack renang, selebrasi kelas, konsumsi kegiatan",
        "Transportasi": "Kompensasi outing perpustakaan, iktikaf, gladi resik",
        "Administrasi": "Biaya admin bank bulanan, biaya transfer",
    }

    sub_masuk  = defaultdict(float)
    sub_keluar = defaultdict(float)

    for t in kas:
        if t["arus_kas"] == "KAS MASUK":
            sub_masuk[t["sub"]]  += t["value"]
        else:
            sub_keluar[t["sub"]] += t["value"]

    pemasukan = [
        {
            "sub":   KAS_MASUK_LABEL.get(s, s),
            "nilai": round(v),
            "desc":  KAS_MASUK_DESC.get(s, ""),
        }
        for s, v in sorted(sub_masuk.items(), key=lambda x: -x[1])
        if v > 0
    ]
    pengeluaran = [
        {
            "sub":   s,
            "nilai": round(v),
            "desc":  KAS_KELUAR_DESC.get(s, ""),
        }
        for s, v in sorted(sub_keluar.items(), key=lambda x: -x[1])
        if v > 0
    ]
    return {"pemasukan": pemasukan, "pengeluaran": pengeluaran}


def compute_kegiatan(txns):
    """Pecahan sub-program untuk kategori Kegiatan."""
    keg = [t for t in txns if t["kategori"] == "Kegiatan"]

    SUB_CONFIG = {
        "Outing":     {"label": "Outing Udjo Bandung",   "icon": "🏕️",  "desc": "Saung Angklung, workshop peuyeum, menginap",           "periode": "Juli 2025"},
        "Research":   {"label": "Research Day",           "icon": "🔬",  "desc": "Presentasi penelitian siswa, pameran ilmiah",          "periode": "Oktober–Nov 2025"},
        "OASIXTH":    {"label": "OASIXTH Trip",           "icon": "⛰️",  "desc": "Perjalanan akhir SD angkatan ke-6",                    "periode": "April–Juni 2026"},
        "Graduation": {"label": "Graduation SD6 Bangga!", "icon": "🎓",  "desc": "Perpisahan, buku tahunan, kenang-kenangan Fasil",      "periode": "Mei–Juni 2026"},
        "Lainnya":    {"label": "Kegiatan Lainnya",       "icon": "🎪",  "desc": "Iktikaf, mutasi internal, kegiatan incidental",        "periode": "Sepanjang tahun"},
    }
    ORDER = ["Outing", "Research", "OASIXTH", "Graduation", "Lainnya"]

    subs = defaultdict(lambda: {"masuk": 0.0, "keluar": 0.0})
    for t in keg:
        if t["arus_kas"] == "KAS MASUK":
            subs[t["sub"]]["masuk"]  += t["value"]
        else:
            subs[t["sub"]]["keluar"] += t["value"]

    result = []
    for key in ORDER:
        if key not in subs:
            continue
        masuk  = round(subs[key]["masuk"])
        keluar = round(subs[key]["keluar"])
        cfg    = SUB_CONFIG.get(key, {"label": key, "icon": "", "desc": "", "periode": ""})
        result.append({
            "sub":     key,
            "label":   cfg["label"],
            "icon":    cfg["icon"],
            "desc":    cfg["desc"],
            "periode": cfg["periode"],
            "masuk":   masuk,
            "keluar":  keluar,
            "saldo":   masuk - keluar,
        })
    return result


def compute_tawaun_penyaluran(txns):
    """Semua transaksi penyaluran ta'awun (KAS KELUAR, kategori Ta'awun)."""
    penyaluran = [
        t for t in txns
        if t["kategori"] == "Ta'awun" and t["arus_kas"] == "KAS KELUAR"
    ]
    return [
        {
            "keterangan": t["keterangan"],
            "nilai":      round(t["value"]),
            "periode":    BULAN_LABEL.get(t["bulan"], t["bulan"]),
        }
        for t in penyaluran
    ]


def compute_sunburst(txns):
    """Data 3-level untuk sunburst: arus_kas → kategori → sub_kategori → nilai."""
    tree = defaultdict(lambda: defaultdict(float))
    for t in txns:
        tree[(t["arus_kas"], t["kategori"])][t["sub"]] += t["value"]

    result = {}
    for (arus, kat), subs in tree.items():
        result.setdefault(arus, {})[kat] = {
            k: round(v) for k, v in subs.items() if v > 0
        }
    return result


# ══════════════════════════════════════════════════════════════════
# 3. MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    if not EXCEL_FILE.exists():
        print(f"ERROR: File tidak ditemukan: {EXCEL_FILE}")
        sys.exit(1)

    print(f"Membaca: {EXCEL_FILE.name} ...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=True)

    txns   = load_cashflow(wb)
    print(f"  {len(txns)} transaksi ditemukan")

    summary            = compute_summary(txns)
    rekap              = compute_rekap(txns)
    monthly            = compute_monthly(txns)
    kas_breakdown      = compute_kas_breakdown(txns)
    kegiatan           = compute_kegiatan(txns)
    tawaun_penyaluran  = compute_tawaun_penyaluran(txns)

    # Terapkan semua override kustom dari highlights_override.json (tidak tertimpa update Excel)
    highlights_file = BASE_DIR / "highlights_override.json"
    if highlights_file.exists():
        overrides = json.loads(highlights_file.read_text(encoding="utf-8"))
        applied_hl = sum(1 for m in monthly if m["bulan"] in overrides)
        for m in monthly:
            if m["bulan"] in overrides:
                m["highlight"] = overrides[m["bulan"]]
        if applied_hl:
            print(f"  Highlight override: {applied_hl} bulan")
        if "tawaunPenyaluran" in overrides:
            cats      = overrides["tawaunPenyaluran"]
            cat_total = [0] * len(cats)
            catchall  = next((i for i, c in enumerate(cats) if not c.get("matchKeywords")), len(cats) - 1)
            for t in tawaun_penyaluran:          # gunakan data aktual Excel
                ket   = t["keterangan"].lower()
                placed = False
                for i, cat in enumerate(cats):
                    kws = [kw.lower() for kw in cat.get("matchKeywords", [])]
                    if kws and any(kw in ket for kw in kws):
                        cat_total[i] += t["nilai"]
                        placed = True
                        break
                if not placed:
                    cat_total[catchall] += t["nilai"]
            tawaun_penyaluran = [
                {"keterangan": c["keterangan"], "nilai": cat_total[i], "periode": c["periode"]}
                for i, c in enumerate(cats)
            ]
            print(f"  Ta'awun penyaluran override: {len(tawaun_penyaluran)} kategori "
                  f"(total Rp{sum(cat_total):,.0f})")
    tawaun_siswa       = load_tawaun_siswa(wb)
    rencana            = load_rencana(wb)
    siswa              = load_siswa(wb)
    total_tabungan     = load_total_tabungan(wb)
    sunburst           = compute_sunburst(txns)

    total_rencana         = sum(r["nilai"] for r in rencana)
    saldo_setelah_rencana = summary["saldoAkhir"] - round(total_rencana)

    # Meta info (update manual jika berubah)
    meta = {
        "periode":      "Juli 2025 — Juni 2026",
        "tahunAjaran":  "2025 / 2026",
        "jumlahSiswa":  len(siswa) if siswa else 34,
        "jumlahKelas":  "Kelas Banyu & Jenggala",
        "bendahara":    "OTS Adli",
        "ketuaDK":      "OTS Nisa",
        "noRek":        "174-001-114-5778",
        "namaRek":      "Wahyuni",
        "tglLaporan":   "Juni 2026",
        "saldoRek":     summary["saldoAkhir"],
        "linkLaporan":  "https://bit.ly/Kas_SD6_Bangga",
    }

    payload = {
        "meta":                meta,
        "summary":             summary,
        "rekap":               rekap,
        "monthly":             monthly,
        "kasBreakdown":        kas_breakdown,
        "kegiatan":            kegiatan,
        "tawaunSiswa":         tawaun_siswa,
        "tawaunPenyaluran":    tawaun_penyaluran,
        "rencana":             rencana,
        "totalRencana":        round(total_rencana),
        "saldoSetelahRencana": round(saldo_setelah_rencana),
        "siswa":               siswa,
        "sunburst":            sunburst,
        "generatedAt":         datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    js = (
        "// Auto-generated oleh generate_report.py — JANGAN edit manual\n"
        f"// Dibuat: {payload['generatedAt']}\n"
        f"window.KASDATA = {json.dumps(payload, ensure_ascii=False, indent=2)};\n"
    )

    OUTPUT_JS.write_text(js, encoding="utf-8")
    print(f"Berhasil: {OUTPUT_JS.name}")
    print(f"  Total Masuk : Rp{summary['totalMasuk']:,.0f}")
    print(f"  Total Keluar: Rp{summary['totalKeluar']:,.0f}")
    print(f"  Saldo Akhir : Rp{summary['saldoAkhir']:,.0f}")

    # Inline data ke HTML agar file HTML bisa dishare tanpa data.js
    html_file = BASE_DIR / "Laporan_Pertanggungjawaban_Pengelolaan_Dana_SD6_SAI_Meruyung.html"
    if html_file.exists():
        html_content = html_file.read_text(encoding="utf-8")
        inline_block = (
            f'<script id="inlineData">\n'
            f'window.KASDATA = {json.dumps(payload, ensure_ascii=False, indent=2)};\n'
            f'</script>'
        )
        updated = re.sub(
            r'<script id="inlineData">.*?</script>',
            inline_block,
            html_content,
            flags=re.DOTALL,
        )
        if updated != html_content:
            html_file.write_text(updated, encoding="utf-8")
            print(f"  Share mode : HTML diperbarui (data inline)")
        elif re.search(r'<script id="inlineData">', html_content):
            print(f"  Share mode : HTML sudah terkini, tidak ada perubahan")
        else:
            print(f"  Share mode : placeholder inlineData tidak ditemukan, skip")
    print()
    print("Buka Laporan_Pertanggungjawaban_Pengelolaan_Dana_SD6_SAI_Meruyung.html di browser.")


if __name__ == "__main__":
    main()
